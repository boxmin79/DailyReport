from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime, date
import xlwings as xw
import os
import win32com.client as win32


def get_new_id(sheet_name):
    wb = load_workbook(r'D:/My Documents/운행일보/dr.xlsm', read_only=True)
    ws = wb[sheet_name]
    new_id = ws.cell(row=ws.max_row, column=1).value + 1

    wb.close()

    return new_id


def input_dicts(info_dict, sheet_name):
    wb = load_workbook(r'D:/My Documents/운행일보/dr.xlsm', keep_vba=True)
    ws = wb[sheet_name]
    ws.append(list(info_dict.values()))

    wb.save(r'D:/My Documents/운행일보/dr.xlsm')
    wb.close()


def print_envelop(print_dict):
    wb = xw.Book(r'D:/My Documents/운행일보/dr.xlsm')
    ws_print = wb.sheets(7)

    for idx in print_dict:
        ws_print["F28"].value = print_dict[idx]['우편물주소']
        ws_print["E28"].value = print_dict[idx]['상호']
        ws_print["C45"].value = print_dict[idx]['우편번호'][0]
        ws_print["C48"].value = print_dict[idx]['우편번호'][1]
        ws_print["C51"].value = print_dict[idx]['우편번호'][2]
        ws_print["C54"].value = print_dict[idx]['우편번호'][3]
        ws_print["C57"].value = print_dict[idx]['우편번호'][4]

        print_macro = wb.macro('print_envelop')
        print_macro()
        exit_macro = wb.macro('Exit_Excel')
        exit_macro()

    # wb.close()


def record_iti_date(print_dict):
    # print(print_dict)
    wb = load_workbook(r'D:/My Documents/운행일보/dr.xlsm', keep_vba=True)
    ws = wb['세금계산서']

    for idx in print_dict:
        for row in ws[2:ws.max_row]:
            if row[0].value == print_dict[idx]['일보ID']:
                row[8].value = datetime.now().date()

    wb.save(r'D:/My Documents/운행일보/dr.xlsm')
    wb.close()


def print_iti(print_dict):
    wb = xw.Book(r'D:/My Documents/운행일보/dr.xlsm')
    ws_print = wb.sheets(6)

    for idx in print_dict:
        fare = str(print_dict[idx]['합계금액'])
        tax = str(int(print_dict[idx]['합계금액'] / 10))
        add = print_dict[idx]['사업장주소'].replace('광역시', '시')
        load_date = datetime.strftime(print_dict[idx]['배차시간'], '%Y년 %m월 %d일').replace(' ', '')

        ws_print["X5"].value = print_dict[idx]['사업자번호']
        ws_print["X6"].value = print_dict[idx]['상호']
        ws_print["AG6"].value = print_dict[idx]['대표자명']
        ws_print["X7"].value = print_dict[idx]['사업장주소'].replace('광역시', '시')
        ws_print["X8"].value = print_dict[idx]['업태']
        ws_print["AE8"].value = print_dict[idx]['업종']
        ws_print["B11"].value = load_date
        ws_print["G11"].value = 11 - len(fare)
        ws_print["H11"].value = f'{fare: >11}'[0]
        ws_print["I11"].value = f'{fare: >11}'[1]
        ws_print["J11"].value = f'{fare: >11}'[2]
        ws_print["K11"].value = f'{fare: >11}'[3]
        ws_print["L11"].value = f'{fare: >11}'[4]
        ws_print["M11"].value = f'{fare: >11}'[5]
        ws_print["N11"].value = f'{fare: >11}'[6]
        ws_print["O11"].value = f'{fare: >11}'[7]
        ws_print["P11"].value = f'{fare: >11}'[8]
        ws_print["Q11"].value = f'{fare: >11}'[9]
        ws_print["R11"].value = f'{fare: >11}'[10]
        ws_print["T11"].value = f'{tax: >10}'[0]
        ws_print["U11"].value = f'{tax: >10}'[1]
        ws_print["V11"].value = f'{tax: >10}'[2]
        ws_print["W11"].value = f'{tax: >10}'[3]
        ws_print["X11"].value = f'{tax: >10}'[4]
        ws_print["Y11"].value = f'{tax: >10}'[5]
        ws_print["Z11"].value = f'{tax: >10}'[6]
        ws_print["AA11"].value = f'{tax: >10}'[7]
        ws_print["AB11"].value = f'{tax: >10}'[8]
        ws_print["AC11"].value = f'{tax: >10}'[9]
        ws_print["B13"].value = load_date[5:]
        ws_print["E13"].value = print_dict[idx]['상차지']
        ws_print["N13"].value = print_dict[idx]['하차지']
        ws_print["X13"].value = print_dict[idx]['합계금액']
        ws_print["AE13"].value = print_dict[idx]['합계금액'] / 10
        ws_print["B18"].value = print_dict[idx]['합계금액'] + (print_dict[idx]['합계금액'] / 10)

        ws_print["X24"].value = print_dict[idx]['사업자번호']
        ws_print["X25"].value = print_dict[idx]['상호']
        ws_print["AG25"].value = print_dict[idx]['대표자명']
        ws_print["X26"].value = print_dict[idx]['사업장주소'].replace('광역시', '시')
        ws_print["X27"].value = print_dict[idx]['업태']
        ws_print["AE27"].value = print_dict[idx]['업종']
        ws_print["B30"].value = load_date
        ws_print["G30"].value = 11 - len(fare)
        ws_print["H30"].value = f'{fare: >11}'[0]
        ws_print["I30"].value = f'{fare: >11}'[1]
        ws_print["J30"].value = f'{fare: >11}'[2]
        ws_print["K30"].value = f'{fare: >11}'[3]
        ws_print["L30"].value = f'{fare: >11}'[4]
        ws_print["M30"].value = f'{fare: >11}'[5]
        ws_print["N30"].value = f'{fare: >11}'[6]
        ws_print["O30"].value = f'{fare: >11}'[7]
        ws_print["P30"].value = f'{fare: >11}'[8]
        ws_print["Q30"].value = f'{fare: >11}'[9]
        ws_print["R30"].value = f'{fare: >11}'[10]
        ws_print["T30"].value = f'{tax: >10}'[0]
        ws_print["U30"].value = f'{tax: >10}'[1]
        ws_print["V30"].value = f'{tax: >10}'[2]
        ws_print["W30"].value = f'{tax: >10}'[3]
        ws_print["X30"].value = f'{tax: >10}'[4]
        ws_print["Y30"].value = f'{tax: >10}'[5]
        ws_print["Z30"].value = f'{tax: >10}'[6]
        ws_print["AA30"].value = f'{tax: >10}'[7]
        ws_print["AB30"].value = f'{tax: >10}'[8]
        ws_print["AC30"].value = f'{tax: >10}'[9]
        ws_print["B32"].value = load_date[5:]
        ws_print["E32"].value = print_dict[idx]['상차지']
        ws_print["N32"].value = print_dict[idx]['하차지']
        ws_print["X32"].value = print_dict[idx]['합계금액']
        ws_print["AE32"].value = print_dict[idx]['합계금액'] / 10
        ws_print["B37"].value = print_dict[idx]['합계금액'] + (print_dict[idx]['합계금액'] / 10)

        print_macro = wb.macro('print_ti')
        print_macro()
        exit_macro = wb.macro('Exit_Excel')
        exit_macro()


def not_deposit_list():
    wb = load_workbook(r'D:/My Documents/운행일보/dr.xlsm')
    ws_iti = wb['세금계산서 발행']
    ws_dr = wb['운행일보']
    ws_csn = wb['운송사목록']

    wb_new = Workbook()
    ws_new = wb_new.active

    no_deposit_list = []
    pn = ''
    for iti_row in ws_iti[2:ws_iti.max_row]:
        if iti_row[10].value is None:
            # print(iti_row)
            for csn_row in ws_csn[2:ws_csn.max_row]:
                if csn_row[0].value == iti_row[2].value:
                    # print(csn_row[8].value)
                    pn = csn_row[8].value
                    # print(phone_num)

            ndl_row = ws_dr[iti_row[0].row]
            # print(ndl_row)

            no_deposit_list.append(
                [ndl_row[0].value, datetime.date(ndl_row[1].value), ndl_row[3].value, ndl_row[4].value,
                 ndl_row[6].value, pn, iti_row[7].value])

    # print(np.array(no_deposit_list))
    ws_new.append(['일보ID', '날짜', '상차지', '하차지', '화주명', '연락처', '합계금액'])
    for ndl in no_deposit_list:
        # print(ndl)
        ws_new.append(ndl)

    wb_new.save(r'D:\My Documents\운행일보\미입금리스트.xlsx')

    wb.close()
    wb_new.close()


def get_max_date():
    wb = load_workbook(r'D:/My Documents/운행일보/dr.xlsm')
    ws = wb['입금내역']
    # ws = wb['입금내역 (2)']

    date_list = []
    max_date = date.fromisoformat('1900-01-01')
    if ws.max_row >= 2:
        for row in ws[2:ws.max_row]:
            # print(f'type({type(row[1].value)}), value({row[1].value})')
            date_list.append(row[1].value.date())
        max_date = max(date_list)
    wb.close()

    return max_date


def get_exception_list():
    wb = load_workbook(r'D:/My Documents/운행일보/dr.xlsm', keep_vba=True)
    ws = wb['예외항목']

    exception_list = []
    for cell in ws["A"]:
        exception_list.append(cell.value)

    wb.close()
    return exception_list


def get_iti_list():
    wb = load_workbook(r'D:/My Documents/운행일보/dr.xlsm')
    ws_iti = wb['세금계산서 발행']
    # key_list = ['iid', '행번호', '일보ID', '날짜', '운송사', '운임', '입금일']
    iti_list = []
    # i = 0
    for row in ws_iti[2:ws_iti.max_row]:
        # print(row)
        if row[10].value is None:
            iti_list.append({'행번호': row[0].row, '일보ID': row[0].value, '날짜': row[1].value.date(), '운송사': row[3].value, '운임': row[7].value,
                             '입금일': row[10].value, '입금명': ''})
            # i += 1
    wb.close()
    # for il in iti_list:
    #     print(il)
    return iti_list


def get_dp_list(file_deposit, max_date, exception_list):
    # print(type(max_date))
    dp_list = []

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    wb = excel.Workbooks.Open(file_deposit)
    ws = wb.ActiveSheet

    # i = 0
    # for row in range(6, 300):
    #     # print(ws.Cells(row, 1).Value)
    #     if ws.Cells(row, 1).Value is not None:
    #         date = datetime.strptime(ws.Cells(row, 1).Value, '%Y.%m.%d %H:%M:%S').date()
    #         if ws.Cells(row, 6).Value > 0 and date > max_date:
    #             name = ws.Cells(row, 3).Value.strip('\xa0')
    #             deposit_amount = int(ws.Cells(row, 6).Value)
    #             if name not in exception_list:
    #                 dp_list.append({'번호': i, '입금일': date, '보낸분': name, '입금액': deposit_amount})
    #                 i += 1
    i = 0
    row = 6
    while ws.Cells(row, 1).Value is not None:

        # for column in range(1,9):
        #     print(ws.Cells(row, column).Value, end=', ')
        # print()

        dp_date = datetime.strptime(ws.Cells(row, 1).Value, '%Y.%m.%d %H:%M:%S').date()
        if ws.Cells(row, 6).Value > 0 and dp_date > max_date:
            if ws.Cells(row, 3).Value is not None:
                name = ws.Cells(row, 3).Value.strip('\xa0')
            else:
                name = ws.Cells(row, 2).Value.strip('\xa0')
            deposit_amount = int(ws.Cells(row, 6).Value)
            if name not in exception_list:
                dp_list.append({'번호': i, '입금일': dp_date, '보낸분': name, '입금액': deposit_amount})
                i += 1
        row += 1
    wb.Close()
    excel.Application.Quit()
    # for dl in dp_list:
    #     print(dl)

    return dp_list


def xls_to_xlsx(file_name):
    # 파일이름 경로 분리하기
    last_slash_index = find_index_last_slash(file_name)
    file_path = file_name[:last_slash_index]
    file_name = file_name[last_slash_index:]

    # xls를 xlsx로 바꾸기
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(file_path+file_name)

    wb.SaveAs(file_name + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
    print(f'<D:/My Documents/{file_name}x>를 생성하였습니다.')
    wb.Close()  # FileFormat = 56 is for .xls extension
    excel.Application.Quit()

    # 'xls'파일 삭제
    if os.path.isfile(file_path+file_name):
        os.remove(file_path+file_name)
        print('<file_path+file_name>를 삭제 하였습니다.')
    print(f'<D:/My Documents/{file_name}x>를 반환 합니다.')
    return 'D:/My Documents/' + file_name + 'x'


def find_index_last_slash(str_to_find):
    result = 0
    idx = 0
    while idx > -1:
        idx = str_to_find.find('/', idx)
        if idx > -1:
            # print(idx)
            idx += len('/')
            result = idx
    return result


def add_exception(exception_name):
    wb = load_workbook(r'D:/My Documents/운행일보/dr.xlsm', keep_vba=True)
    ws = wb['예외항목']
    ws.cell(row=ws.max_row + 1, column=1).value = exception_name

    wb.save(r'D:/My Documents/운행일보/dr.xlsm')
    wb.close()


def input_match_data(mat_dict):
    wb = load_workbook(r'D:/My Documents/운행일보/dr.xlsm', keep_vba=True)
    ws_iti = wb['세금계산서']
    ws_dp = wb['입금내역']

    new_id = ws_dp.cell(row=ws_dp.max_row, column=1).value + 1
    # # 1. '일보ID'로 row 가져오기
    for idx in mat_dict:
        # 세금계산서 시트 입력
        if str(type(idx)) == "<class 'int'>":
            print(f'type <int>  {idx}: {mat_dict[idx]}')
            for row in ws_iti[2:ws_iti.max_row]:
                if row[0].value == mat_dict[idx]['일보ID']:
                    row[10].value = mat_dict[idx]['입금일'].date()
                    row[11].value = mat_dict[idx]['보낸분']
        if str(type(idx)) == "<class 'str'>":
            if '중복' in idx:
                print(f'type <str> 중복 {idx}: {mat_dict[idx]}')
            else:
                print(f'type <str> {idx}: {mat_dict[idx]}')

        # 입금내역 시트 입력
        ws_dp.append([new_id, mat_dict[idx]['입금일'].date(), mat_dict[idx]['보낸분'], mat_dict[idx]['입금액'],
                      mat_dict[idx]['입금일'], mat_dict[idx]['일보ID']])
        new_id += 1

    wb.save(r'D:/My Documents/운행일보/dr.xlsm')
    wb.close()


def input_cost_dict_to_excel(cost_dict):
    wb = load_workbook(r'D:/My Documents/운행일보/dr.xlsm', keep_vba=True)
    ws = wb['비용']
    ws.append(list(cost_dict.values()))
    # print(list(cost_dict.values()))
    # for idx in cost_dict:
    #     print(f'{idx}: {cost_dict[idx]}')

    wb.save(r'D:/My Documents/운행일보/dr.xlsm')
    wb.close()


def insert_csn_record_in_excel(csn_dict):
    print('Run <ExcelMod.insert_csn_record_in_excel()>')
    wb = load_workbook(r'D:/My Documents/운행일보/dr.xlsm', keep_vba=True)
    # print(wb.sheetnames)
    ws_csn = wb['운송사목록']

    ws_csn.append(list(csn_dict.values()))
    # print(new_csn_record)

    wb.save(r'D:/My Documents/운행일보/dr.xlsm')
    wb.close()
    print('Done <ExcelMod.insert_csn_record_in_excel()>')


def get_row_by_csn_id(csn_num):
    wb = load_workbook(r'D:/My Documents/운행일보/dr.xlsm', read_only=True)
    ws = wb['운송사목록']
    result = None
    for row in ws[2:ws.max_row]:
        if row[0].value == csn_num:
            result = row[0].row
    wb.close()
    return result


def edit_csn_record(row, csn_dict):
    print('<ExcelMod.edit_csn_record> 실행')
    wb = load_workbook(r'D:/My Documents/운행일보/dr.xlsm', keep_vba=True)
    ws = wb['운송사목록']
    print('\t\t [csn_dict]를 리스트로 변환 합니다.')
    csn_list = list(csn_dict.values())
    print(f'\t\t csn_list : {csn_list}')
    print('\t\t리스트를 셀에 입력합니다.')
    i = 0
    for cell in ws[row]:
        cell.value = csn_list[i]
        i += 1
    print('\t\t셀에 리스트 입력 완료')
    wb.save(r'D:/My Documents/운행일보/dr.xlsm')
    wb.close()
    print('<ExcelMod.edit_csn_record> 종료')


def delete_row(row_num):
    wb = load_workbook(r'D:/My Documents/운행일보/dr.xlsm', keep_vba=True)
    ws = wb['운송사목록']
    ws.delete_rows(row_num)
    wb.save(r'D:/My Documents/운행일보/dr.xlsm')
    wb.close()
